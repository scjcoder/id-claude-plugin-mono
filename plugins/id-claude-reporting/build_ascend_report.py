#!/usr/bin/env python3
"""
Build the Ascend API Monthly Activity Report for May 2026.
Three sections: Active Full Month | Newly Onboarded | Offboarded Still Billed
"""

import json, subprocess, sys, os
from datetime import datetime
from collections import defaultdict

OUTDIR   = "/Users/sean/CODE/id-claude-reporting"
SLUG     = "2026-05"
MONTH    = "May 2026"
GEN_PY   = "/Users/sean/CODE/id-claude-reporting/skills/ascend-activity-report/generate_report.py"

# ── Load data ──────────────────────────────────────────────────────────────────
with open(f"{OUTDIR}/ascend_facilities_2026-05.json") as f:
    baseline = json.load(f)   # [{facilityId, clientName, facilityName, status}]

with open(f"{OUTDIR}/ascend_snapshots_2026-05.json") as f:
    snaps_raw = json.load(f)  # dict keyed by facilityId string

with open(f"{OUTDIR}/ascend_april_ids.json") as f:
    april_data = json.load(f)

snaps      = {str(k): v for k, v in snaps_raw.items()}
april_ids  = set(april_data["aprilIds"])
new_in_may = set(april_data["newInMayIds"])  # in May but not April = newly onboarded

# Inactive baseline IDs (cancelled but had May snapshots)
inactive_baseline_ids = {
    str(f["facilityId"]) for f in baseline if f.get("status") == "inactive_had_may_snapshot"
}

# ── Categorise each baseline facility ─────────────────────────────────────────
active_locations     = []
onboarded_locations  = []
offboarded_locations = []
inactive_locations   = []  # zero snapshots this month

def fmt_client(name):
    return name.replace("-", " ")

for fac in baseline:
    fid   = str(fac["facilityId"])
    cname = fmt_client(fac["clientName"])
    fname = fac["facilityName"]

    snap = snaps.get(fid)
    if snap:
        last_fmt = datetime.strptime(snap["lastReceived"], "%Y-%m-%d").strftime("%b %d, %Y")
        entry = {
            "facility_id":    fid,
            "client":         cname,
            "location":       fname,
            "snapshot_count": snap["count"],
            "last_snapshot":  last_fmt,
        }
        if fid in inactive_baseline_ids:
            offboarded_locations.append(entry)
        elif fid in new_in_may:
            onboarded_locations.append(entry)
        else:
            active_locations.append(entry)
    else:
        inactive_locations.append({"facility_id": fid, "client": cname, "location": fname})

for lst in (active_locations, onboarded_locations, offboarded_locations, inactive_locations):
    lst.sort(key=lambda x: (x["client"], x["location"]))

print(f"Active full month: {len(active_locations)}")
print(f"Newly onboarded:   {len(onboarded_locations)}")
print(f"Offboarded billed: {len(offboarded_locations)}")
print(f"Inactive (0 snaps):{len(inactive_locations)}")

# ── Excel ─────────────────────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl",
                    "--break-system-packages", "-q"], check=True)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "All Locations"

NAVY_FILL   = PatternFill("solid", fgColor="1B2B4B")
GREEN_FILL  = PatternFill("solid", fgColor="E8F5EE")
TEAL_FILL   = PatternFill("solid", fgColor="E8F4F8")
AMBER_FILL  = PatternFill("solid", fgColor="FEF9E7")
RED_FILL    = PatternFill("solid", fgColor="FDECEA")
DIV_GREEN   = PatternFill("solid", fgColor="2E7D52")
DIV_TEAL    = PatternFill("solid", fgColor="2E7D9B")
DIV_AMBER   = PatternFill("solid", fgColor="D68910")
DIV_RED     = PatternFill("solid", fgColor="C0392B")
WHITE_FONT  = Font(color="FFFFFF", bold=True, name="Calibri")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
BODY_FONT   = Font(name="Calibri", size=10)
thin        = Side(style="thin", color="D5D8DC")
BORDER      = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=True)

headers = ["GoldenEye ID", "Client", "Location", "Snapshot Count", "Last Snapshot Date", "Status"]
ws.append(headers)
for col in range(1, 7):
    cell = ws.cell(1, col)
    cell.fill = NAVY_FILL; cell.font = HEADER_FONT
    cell.alignment = CENTER; cell.border = BORDER
ws.row_dimensions[1].height = 30
ws.freeze_panes = "A2"

def write_divider(ws, label, fill):
    ws.append([label])
    row = ws.max_row
    ws.merge_cells(f"A{row}:F{row}")
    cell = ws.cell(row, 1)
    cell.fill = fill; cell.font = WHITE_FONT; cell.alignment = LEFT
    ws.row_dimensions[row].height = 22

def write_row(ws, fid, client, location, count, last, status, row_fill):
    ws.append([fid, client, location,
               count if count is not None else "—",
               last  if last  else "—", status])
    row = ws.max_row
    ws.row_dimensions[row].height = 28
    for col in range(1, 7):
        cell = ws.cell(row, col)
        cell.border = BORDER; cell.font = BODY_FONT
        cell.alignment = CENTER if col in (1, 4) else LEFT
        cell.fill = row_fill

sections = [
    ("Newly Onboarded This Month", onboarded_locations,  DIV_TEAL,  TEAL_FILL,  "Newly Onboarded"),
    ("Offboarded — Still Billed",  offboarded_locations, DIV_AMBER, AMBER_FILL, "Offboarded – Still Billed"),
    ("Active — Full Month",        active_locations,     DIV_GREEN, GREEN_FILL, "Active"),
    ("No Snapshots This Month",    inactive_locations,   DIV_RED,   RED_FILL,   "Inactive – No Snapshots"),
]
for label, locs, div_fill, row_fill, status_label in sections:
    if not locs:
        continue
    write_divider(ws, f"{label} ({len(locs)})", div_fill)
    for loc in locs:
        write_row(ws, loc["facility_id"], loc["client"], loc["location"],
                  loc.get("snapshot_count"), loc.get("last_snapshot"), status_label, row_fill)

col_widths = [14, 32, 42, 16, 22, 26]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── Sheet 2: Client Summary ────────────────────────────────────────────────────
ws2 = wb.create_sheet("Client Summary")
client_stats = defaultdict(lambda: {"active": 0, "onboarded": 0, "offboarded": 0, "inactive": 0})
for loc in active_locations:     client_stats[loc["client"]]["active"]     += 1
for loc in onboarded_locations:  client_stats[loc["client"]]["onboarded"]  += 1
for loc in offboarded_locations: client_stats[loc["client"]]["offboarded"] += 1
for loc in inactive_locations:   client_stats[loc["client"]]["inactive"]   += 1

summary_rows = []
for client, s in client_stats.items():
    total_c = s["active"] + s["onboarded"] + s["offboarded"] + s["inactive"]
    billed  = s["active"] + s["onboarded"] + s["offboarded"]
    summary_rows.append((client, total_c, s["active"], s["onboarded"], s["offboarded"], s["inactive"]))
summary_rows.sort(key=lambda x: x[0])

ws2.append(["Client", "Total Billed", "Active Full Month", "Newly Onboarded", "Offboarded Billed", "No Snapshots"])
for col in range(1, 7):
    cell = ws2.cell(1, col)
    cell.fill = NAVY_FILL; cell.font = HEADER_FONT
    cell.alignment = CENTER; cell.border = BORDER
ws2.row_dimensions[1].height = 30
ws2.freeze_panes = "A2"

for row_data in summary_rows:
    ws2.append(list(row_data))
    row = ws2.max_row
    ws2.row_dimensions[row].height = 26
    for col in range(1, 7):
        cell = ws2.cell(row, col)
        cell.border = BORDER; cell.font = BODY_FONT
        cell.alignment = LEFT if col == 1 else CENTER

col_widths2 = [36, 14, 20, 18, 20, 14]
for i, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

xlsx_path = f"{OUTDIR}/ascend_activity_{SLUG}.xlsx"
wb.save(xlsx_path)
print(f"Excel saved: {xlsx_path}")

# ── PDF ───────────────────────────────────────────────────────────────────────
report_data = {
    "month_label":       MONTH,
    "total_locations":   len(baseline),
    "active_count":      len(active_locations),
    "onboarded_count":   len(onboarded_locations),
    "offboarded_count":  len(offboarded_locations),
    "inactive_count":    len(inactive_locations),
    "active_locations":     active_locations,
    "onboarded_locations":  onboarded_locations,
    "offboarded_locations": offboarded_locations,
    "inactive_locations":   inactive_locations,
}

result = subprocess.run(
    [sys.executable, GEN_PY,
     "--month", MONTH, "--slug", SLUG, "--outdir", OUTDIR,
     "--data", json.dumps(report_data)],
    capture_output=True, text=True
)
print(result.stdout)
if result.returncode != 0:
    print("PDF ERROR:", result.stderr, file=sys.stderr); sys.exit(1)

pdf_path = f"{OUTDIR}/ascend_activity_{SLUG}.pdf"
print(f"PDF ready: {pdf_path}  ({os.path.getsize(pdf_path)/1024:.1f} KB)")
print("DONE")
